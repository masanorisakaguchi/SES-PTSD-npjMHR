"""★2026-08-16 廃止（記録として残す。新しい計算にこのファイルを使わないこと）。

このスクリプトには日跨ぎのバグがある。104行 `onset = min(on)` と 117行の3分割が
「その日の何秒目か」で並べているため、音が日を跨ぐ H08 では原点として
00:39:04 を拾う（真の開始は 23:58:28）。基準90秒が刺激中に入り込み、相関が −0.95 に化けた。

正本は `ses_scl_absolute_time_20260816.py`（絶対時刻・セッション別の再生カウンタ）。
本文・回答書の SCL の値はすべてそちらから出ている。
監査記録は `../16_2214_kaiseki_san_日跨ぎ監査_7本の判定結果.md`、
型は解析さんの簿 P-001 / P-027（`python3 _tools/kaiseki/pitfall.py brief`）。

--- 以下、当時の説明（原文のまま）---

SCL over the FULL 600-s sound-delivery window (Reviewer 1).

Figure S2 legend: "SCL values were standardized within participants using the 90-s pre-epoch
baseline (z = (SCL - baseline mean) / baseline SD)".  n = 5; H03 excluded (SCL recording failure).

Validation: reproduce the published per-third values before trusting the full-window value.
"""
import csv, io, os, sys, math, statistics as st

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
SES = os.environ.get("SES_STAGE_ROOT", "./_private/stage_scoring")

PIDS = ["H01", "H02", "H08", "H09", "H14"]        # H03 excluded (no SCL)
SUDS = {"H01": 0.1667, "H02": 0.1429, "H08": -0.5, "H09": -0.8, "H14": -0.2}
PUB = {"H01": (-0.3155, 0.3416, 1.5154), "H02": (-5.9145, 0.8383, 16.8802),
       "H08": (13.3831, 5.1388, 4.0298), "H09": (2.1822, 14.5133, 4.0163),
       "H14": (-0.8133, -3.4297, -3.9390)}
EDA = {
    "H01": r"Participants\H01\Visit-4 20241026\EmpaticaData\eda_H01_Visit-4.xlsx",
    "H02": r"Participants\H02\Visit-4\EmpaticaData\eda_H02_Visit-4.csv",
    "H08": r"Participants\H08\Visit-4\EmpaticaData\eda_H08_Visit-4.csv",
    "H09": r"Participants\H09\Visit-4\EmpaticaData\eda_H09_Visit-4.csv",
    "H14": r"Participants\H14\Visit-4\EmpaticaData\eda_H14_Visit-4.csv",
}
SOUND = {
    "H01": [r"Participants\H01\20241026SES-A1\sound_log\records_2024-10-26-21-46.csv"],
    "H02": [r"Participants\未整理データ\UnidentifiedSoundLogs\records_2024-12-21-22-09.csv"],
    "H08": [r"Participants\H08\Visit-4\Soundlog\records_2025-05-23-22-00.csv"],
    "H09": [r"Participants\H09\Visit-4\Soundlog\records_2025-05-23-21-27.csv"],
    "H14": [r"Participants\H14\Visit-4\SoundPlayerLogforSleep\records_2025-08-17-21-35.csv",
            r"Participants\H14\Visit-4\SoundPlayerLogforSleep\records_2025-08-17-23-03.csv",
            r"Participants\H14\Visit-4\SoundPlayerLogforSleep\records_2025-08-18-01-53.csv"],
}


def sound_on(pid):
    out = []
    for rel in SOUND[pid]:
        p = os.path.join(SES, rel)
        if not os.path.exists(p): continue
        with open(p, encoding="utf-8", errors="replace", newline="") as fh:
            rd = csv.reader(fh); next(rd, None); cols = next(rd, None); n = len(cols)
            for r in rd:
                if n == 7 and len(r) >= 7: ts, vol = r[0], r[6]
                elif len(r) >= 8: ts, vol = r[0], r[7]
                else: continue
                try: v = int(float(vol))
                except Exception: continue
                if v <= 0: continue
                try: hh, mm, ss = [int(x) for x in ts.strip().split(" ", 1)[1].split(":")]
                except Exception: continue
                out.append(hh*3600+mm*60+ss)
    return out


def load_eda(pid):
    """-> [(sec_of_day, eda)]"""
    p = os.path.join(SES, EDA[pid])
    out = []
    if p.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            if len(row) < 3 or row[1] is None or row[2] is None: continue
            t = str(row[1])
            try:
                hhmmss = t.split(" ")[1].split("+")[0].split(".")[0]
                hh, mm, ss = [int(x) for x in hhmmss.split(":")]
                out.append((hh*3600+mm*60+ss, float(row[2])))
            except Exception:
                continue
        return out
    with open(p, encoding="utf-8", errors="replace", newline="") as fh:
        rd = csv.reader(fh); next(rd, None)
        for r in rd:
            if len(r) < 3: continue
            try:
                hhmmss = r[1].strip().split(" ")[1].split("+")[0].split(".")[0]
                hh, mm, ss = [int(x) for x in hhmmss.split(":")]
                out.append((hh*3600+mm*60+ss, float(r[2])))
            except Exception:
                continue
    return out


def pearson(x, y):
    n = len(x); mx, my = sum(x)/n, sum(y)/n
    sx = math.sqrt(sum((a-mx)**2 for a in x)); sy = math.sqrt(sum((b-my)**2 for b in y))
    if sx == 0 or sy == 0: return None
    return sum((a-mx)*(b-my) for a, b in zip(x, y))/(sx*sy)


print("=" * 100)
print("SCL, standardised to the 90-s pre-onset baseline, over the sound-delivery window")
print("=" * 100)
print(f"{'PID':<5} {'n samples':>9} {'thirds computed':<34} {'published thirds':<34} {'FULL 600s':>10}")
full = {}
for pid in PIDS:
    on = sound_on(pid)
    if not on:
        print(f"{pid:<5} (no sound log)"); continue
    onset = min(on)
    eda = load_eda(pid)
    base = [v for t, v in eda if onset-90 <= t < onset]
    if len(base) < 5:
        print(f"{pid:<5} baseline samples={len(base)}  -> insufficient"); continue
    bm, bs = st.mean(base), (st.stdev(base) if len(base) > 1 else 0.0)
    if bs == 0:
        print(f"{pid:<5} baseline SD = 0 -> cannot standardise"); continue
    onset_set = set(on)
    zs = [((v-bm)/bs, t) for t, v in eda if t in onset_set]
    if not zs:
        print(f"{pid:<5} no EDA inside sound window"); continue
    # thirds by ordered sound seconds
    ordered = sorted(set(on))
    k = len(ordered)//3
    seg = [set(ordered[:k]), set(ordered[k:2*k]), set(ordered[2*k:])]
    th = []
    for s in seg:
        v = [z for z, t in zs if t in s]
        th.append(round(st.mean(v), 4) if v else None)
    fv = round(st.mean([z for z, _ in zs]), 4)
    full[pid] = fv
    print(f"{pid:<5} {len(zs):>9} {str(tuple(th)):<34} {str(PUB[pid]):<34} {fv:>10}")

ok = [p for p in PIDS if p in full]
if len(ok) >= 3:
    r = pearson([SUDS[p] for p in ok], [full[p] for p in ok])
    print(f"\nPearson r (SUDS change vs mean standardised SCL, full window, n={len(ok)}): "
          f"{'undefined' if r is None else format(r, '+.3f')}")
    print("published per-window r: first -0.5984 / second -0.8005 / third +0.2810 (n=5)")
